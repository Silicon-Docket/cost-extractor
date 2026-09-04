"""Builds the .xlsx report (Summary, Details, and Revisions sheets) from a PipelineResult."""

from __future__ import annotations

from pathlib import Path
from typing import Optional

from openpyxl import Workbook

from cost_extractor.pipeline import PipelineResult
from cost_extractor.revisions import format_revision_timestamp
from cost_extractor import category_rules

_SUMMARY_HEADER = [
    "Document",
    "Status",
    "Amounts Found",
    "Subtotal",
    "Message",
    "Review",
]
_DETAILS_HEADER = [
    "Source File",
    "Location",
    "Matched Text",
    "Rule",
    "Value",
    "Source",
    "Confidence",
    "Review",
    # What OCR originally produced, kept beside the corrected value so a
    # correction reads as a correction rather than a silent rewrite.
    "Read As Text",
    "Category",
    "Category Review",
]

# Written where a value would otherwise be silently trustworthy-looking.
REVIEW_FLAG = "REVIEW"


def review_label(match) -> Optional[str]:
    """What to say about one amount's trustworthiness, in one place.

    Shared with the GUI so the app and the spreadsheet can never disagree
    about whether something counts as checked, corrected, or doubtful.
    None means "nothing worth saying"; a caller rendering into a table cell
    turns that into a blank.

    This is a *current-state* label: a match corrected twice that ends
    back at its original value reads "checked", same as one nobody ever
    touched a second time. That's intentional — this answers "does the
    number differ from the machine reading right now"; the Revisions
    sheet answers "what happened, in order", which is a different
    question.
    """
    if match.value_reviewed:
        return "corrected" if match.effective_value != match.value else "checked"
    return REVIEW_FLAG if match.value_needs_review else None


def category_label(match, rules: list["CategoryRule"]) -> str:
    if match.category_reviewed:
        return match.effective_category
    suggestion = category_rules.suggest_category(match.line_text, rules)
    return f"{suggestion} (suggested, unconfirmed)" if suggestion else "Uncategorized"


def _as_number(value) -> float:
    # Written as a plain float, never an Excel formula string, so a
    # reload via openpyxl (which does not evaluate formulas) sees the
    # real computed value.
    return float(value)


_REVISIONS_HEADER = [
    "Source File",
    "Location",
    "Matched Text",
    "Rule",
    "Revised From",
    "Revised To",
    "Timestamp",
    "Note",
]


def _revision_rows(match) -> list[list]:
    """One row per revision event for one match, in order.

    "Revised From" is the value immediately before that revision: the
    match's original reading for the first revision, the previous
    revision's value for every one after — so reading down a match's
    rows reconstructs the full chain.
    """
    rows = []
    previous = match.value
    for revision in match.value_revisions:
        rows.append(
            [
                match.display_name,
                match.location,
                match.raw_text,
                match.rule_id,
                _as_number(previous),
                _as_number(revision.value),
                format_revision_timestamp(revision.at),
                revision.note,
            ]
        )
        previous = revision.value
    return rows


def build_workbook(
    result: PipelineResult, category_rules: Optional[list["CategoryRule"]] = None
) -> Workbook:
    active_category_rules = category_rules or []
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Summary"
    summary_ws.append(_SUMMARY_HEADER)

    for doc in result.documents:
        summary_ws.append(
            [
                doc.display_name,
                doc.status.value,
                len(doc.matches),
                _as_number(doc.effective_subtotal),
                doc.message,
                REVIEW_FLAG if doc.needs_review else None,
            ]
        )

    # Grand Total keeps meaning the whole batch. The two lines under it say
    # how much of that number rests on a doubtful reading, rather than
    # quietly redefining the headline figure.
    summary_ws.append(
        ["Grand Total", None, None, _as_number(result.effective_grand_total), None]
    )
    summary_ws.append(
        ["Of which needs review", None, None, _as_number(result.review_total), None]
    )
    summary_ws.append(
        ["Confidently read", None, None, _as_number(result.confident_total), None]
    )
    # Counts every unchecked guess, not just low-confidence ones: OCR read
    # $940.00 as $440.00 at 84% confidence, so a score cannot certify a
    # reading as safe.
    summary_ws.append(
        [
            "Guessed amounts not yet checked",
            None,
            None,
            result.unreviewed_ocr_count,
            None,
        ]
    )
    summary_ws.append(
        [
            "Amounts not yet categorized",
            None,
            None,
            result.uncategorized_count,
            None,
        ]
    )

    details_ws = wb.create_sheet("Details")
    details_ws.append(_DETAILS_HEADER)
    for doc in result.documents:
        for m in doc.matches:
            details_ws.append(
                [
                    m.display_name,
                    m.location,
                    m.raw_text,
                    m.rule_id,
                    _as_number(m.effective_value),
                    m.provenance,
                    m.confidence,
                    review_label(m),
                    m.raw_text if m.value_reviewed else None,
                    category_label(m, active_category_rules),
                    REVIEW_FLAG if not m.category_reviewed else None,
                ]
            )

    revisions_ws = wb.create_sheet("Revisions")
    revisions_ws.append(_REVISIONS_HEADER)
    for doc in result.documents:
        for m in doc.matches:
            for row in _revision_rows(m):
                revisions_ws.append(row)

    return wb


def save_workbook(wb: Workbook, path: Path) -> None:
    wb.save(path)
