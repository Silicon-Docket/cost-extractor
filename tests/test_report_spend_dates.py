"""Spend-date columns, Summary row, Revisions Dimension rows, and the
Spend By Month sheet."""

from datetime import date, datetime, timezone
from decimal import Decimal

import openpyxl

from cost_extractor.date_rules import default_rules as default_date_rules
from cost_extractor.extractors.base import Status
from cost_extractor.pipeline import DocumentResult, MatchRecord, PipelineResult
from cost_extractor.report import build_workbook, save_workbook
from cost_extractor.revisions import record_revision

_NOW = datetime(2026, 9, 3, 10, 22, tzinfo=timezone.utc)


def _match(value="100.00", doc_offset=0) -> MatchRecord:
    return MatchRecord(
        display_name="scan.pdf",
        location="page 1",
        raw_text=f"${value}",
        rule_id="standard",
        value=Decimal(value),
        doc_offset=doc_offset,
    )


def _result(matches, full_text="") -> PipelineResult:
    return PipelineResult.from_documents(
        [
            DocumentResult(
                display_name="scan.pdf",
                status=Status.OK,
                matches=matches,
                subtotal=sum((m.value for m in matches), Decimal("0")),
                full_text=full_text,
            )
        ]
    )


def _sheet(tmp_path, result, name, rules=None):
    path = tmp_path / "report.xlsx"
    save_workbook(build_workbook(result, rules), path)
    return openpyxl.load_workbook(path)[name]


def test_build_workbook_with_no_date_rules_argument_still_produces_details(tmp_path):
    # The default-None path -- every pre-existing single-argument call
    # site keeps compiling and behaving as before.
    ws = _sheet(tmp_path, _result([_match()]), "Details")
    header = [c.value for c in ws[1]]
    row = [c.value for c in ws[2]]

    assert row[header.index("Spend Date")] == "Undated"
    assert row[header.index("Spend Date Review")] == "REVIEW"


def test_details_reports_a_confirmed_date(tmp_path):
    m = _match()
    record_revision(m.spend_date_revisions, date(2026, 6, 14), now=_NOW)
    ws = _sheet(tmp_path, _result([m]), "Details")
    header = [c.value for c in ws[1]]
    row = [c.value for c in ws[2]]

    assert row[header.index("Spend Date")] == "2026-06-14"
    assert row[header.index("Spend Date Review")] is None


def test_details_reports_a_confirmed_no_date(tmp_path):
    m = _match()
    record_revision(m.spend_date_revisions, None, now=_NOW)
    ws = _sheet(tmp_path, _result([m]), "Details")
    header = [c.value for c in ws[1]]
    row = [c.value for c in ws[2]]

    assert row[header.index("Spend Date")] == "No Date (confirmed)"
    assert row[header.index("Spend Date Review")] is None


def test_details_reports_a_suggested_unconfirmed_date(tmp_path):
    full_text = "Dated 06/14/2026, amount $100.00."
    m = _match(doc_offset=full_text.index("$100.00"))
    ws = _sheet(
        tmp_path, _result([m], full_text=full_text), "Details", rules=default_date_rules()
    )
    header = [c.value for c in ws[1]]
    row = [c.value for c in ws[2]]

    assert row[header.index("Spend Date")] == "2026-06-14 (suggested, unconfirmed)"
    assert row[header.index("Spend Date Review")] == "REVIEW"


def test_summary_reports_dates_not_yet_reviewed(tmp_path):
    reviewed = _match()
    record_revision(reviewed.spend_date_revisions, date(2026, 6, 14), now=_NOW)
    unreviewed = _match()
    ws = _sheet(tmp_path, _result([reviewed, unreviewed]), "Summary")
    labels = {row[0].value: row[3].value for row in ws.iter_rows()}

    assert labels["Dates Not Yet Reviewed"] == 1


def test_revisions_sheet_gets_a_spend_date_dimension_row(tmp_path):
    m = _match()
    record_revision(m.spend_date_revisions, date(2026, 6, 14), note="from invoice", now=_NOW)
    ws = _sheet(tmp_path, _result([m]), "Revisions")
    header = [c.value for c in ws[1]]
    row = [c.value for c in ws[2]]

    assert row[header.index("Dimension")] == "Spend Date"
    assert row[header.index("Revised From")] == "Undated"
    assert row[header.index("Revised To")] == "2026-06-14"
    assert row[header.index("Note")] == "from invoice"


def test_a_value_revision_row_reads_value_for_dimension(tmp_path):
    m = _match()
    record_revision(m.value_revisions, Decimal("150.00"), now=_NOW)
    ws = _sheet(tmp_path, _result([m]), "Revisions")
    header = [c.value for c in ws[1]]
    row = [c.value for c in ws[2]]

    assert row[header.index("Dimension")] == "Value"


def test_a_second_spend_date_confirmation_shows_two_revision_rows(tmp_path):
    m = _match()
    first = datetime(2026, 9, 3, 10, 14, tzinfo=timezone.utc)
    second = datetime(2026, 9, 3, 10, 22, tzinfo=timezone.utc)
    record_revision(m.spend_date_revisions, date(2026, 6, 1), now=first)
    record_revision(m.spend_date_revisions, date(2026, 6, 14), note="fixed", now=second)
    ws = _sheet(tmp_path, _result([m]), "Revisions")
    header = [c.value for c in ws[1]]
    row1 = [c.value for c in ws[2]]
    row2 = [c.value for c in ws[3]]

    assert row1[header.index("Revised From")] == "Undated"
    assert row1[header.index("Revised To")] == "2026-06-01"
    assert row2[header.index("Revised From")] == "2026-06-01"
    assert row2[header.index("Revised To")] == "2026-06-14"


def test_spend_by_month_sums_matches_into_the_right_month(tmp_path):
    a = _match(value="100.00")
    record_revision(a.spend_date_revisions, date(2026, 6, 14), now=_NOW)
    b = _match(value="50.00")
    record_revision(b.spend_date_revisions, date(2026, 6, 20), now=_NOW)
    c = _match(value="75.00")
    record_revision(c.spend_date_revisions, date(2026, 7, 1), now=_NOW)
    ws = _sheet(tmp_path, _result([a, b, c]), "Spend By Month")
    rows = {row[0]: (row[1], row[2]) for row in ws.iter_rows(min_row=2, values_only=True)}

    assert rows["2026-06"] == (150.0, 2)
    assert rows["2026-07"] == (75.0, 1)


def test_spend_by_month_sorts_chronologically_regardless_of_insertion_order(tmp_path):
    later = _match(value="10.00")
    record_revision(later.spend_date_revisions, date(2026, 8, 1), now=_NOW)
    earlier = _match(value="20.00")
    record_revision(earlier.spend_date_revisions, date(2026, 1, 1), now=_NOW)
    # Constructed in "later, earlier" order to prove the sheet sorts,
    # rather than reflecting incidental dict/insertion order.
    ws = _sheet(tmp_path, _result([later, earlier]), "Spend By Month")
    months = [row[0] for row in ws.iter_rows(min_row=2, max_row=3, values_only=True)]

    assert months == ["2026-01", "2026-08"]


def test_spend_by_month_confirmed_no_date_row(tmp_path):
    m = _match(value="30.00")
    record_revision(m.spend_date_revisions, None, now=_NOW)
    ws = _sheet(tmp_path, _result([m]), "Spend By Month")
    rows = {row[0]: (row[1], row[2]) for row in ws.iter_rows(min_row=2, values_only=True)}

    assert rows["No Date (confirmed)"] == (30.0, 1)
    assert "Not Yet Reviewed" not in rows


def test_spend_by_month_not_yet_reviewed_row(tmp_path):
    m = _match(value="40.00")
    ws = _sheet(tmp_path, _result([m]), "Spend By Month")
    rows = {row[0]: (row[1], row[2]) for row in ws.iter_rows(min_row=2, values_only=True)}

    assert rows["Not Yet Reviewed"] == (40.0, 1)
    assert "No Date (confirmed)" not in rows


def test_a_confirmed_no_date_match_is_distinct_from_an_unreviewed_one(tmp_path):
    # The two-bucket distinction, directly tested: a deliberate "none" and
    # a merely-unreviewed match must never land in the same bucket.
    declined = _match(value="10.00")
    record_revision(declined.spend_date_revisions, None, now=_NOW)
    untouched = _match(value="20.00")
    ws = _sheet(tmp_path, _result([declined, untouched]), "Spend By Month")
    rows = {row[0]: (row[1], row[2]) for row in ws.iter_rows(min_row=2, values_only=True)}

    assert rows["No Date (confirmed)"] == (10.0, 1)
    assert rows["Not Yet Reviewed"] == (20.0, 1)


def test_an_unconfirmed_suggested_date_does_not_reach_a_monthly_total(tmp_path):
    # A rule-suggested-but-unconfirmed date must not put money in a
    # specific month based on a machine guess nobody checked.
    full_text = "Dated 06/14/2026, amount $100.00."
    m = _match(value="100.00", doc_offset=full_text.index("$100.00"))
    ws = _sheet(
        tmp_path,
        _result([m], full_text=full_text),
        "Spend By Month",
        rules=default_date_rules(),
    )
    rows = {row[0]: (row[1], row[2]) for row in ws.iter_rows(min_row=2, values_only=True)}

    assert "2026-06" not in rows
    assert rows["Not Yet Reviewed"] == (100.0, 1)


def test_spend_by_month_buckets_sum_to_the_effective_grand_total(tmp_path):
    dated = _match(value="100.00")
    record_revision(dated.spend_date_revisions, date(2026, 6, 14), now=_NOW)
    declined = _match(value="10.00")
    record_revision(declined.spend_date_revisions, None, now=_NOW)
    untouched = _match(value="20.00")
    result = _result([dated, declined, untouched])
    ws = _sheet(tmp_path, result, "Spend By Month")

    total = sum(row[1] for row in ws.iter_rows(min_row=2, values_only=True))
    assert Decimal(str(total)) == result.effective_grand_total


def test_build_workbook_gains_the_spend_by_month_sheet(tmp_path):
    wb = build_workbook(_result([]))

    assert wb.sheetnames == ["Summary", "Details", "Revisions", "Spend By Month"]
