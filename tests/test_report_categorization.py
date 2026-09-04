"""Category columns, Summary row, Revisions Dimension rows, and the
Categories sheet."""

from datetime import datetime, timezone
from decimal import Decimal

import openpyxl

from cost_extractor.extractors.base import Status
from cost_extractor.pipeline import DocumentResult, MatchRecord, PipelineResult
from cost_extractor.report import build_workbook, save_workbook
from cost_extractor.revisions import record_revision

_NOW = datetime(2026, 9, 3, 10, 22, tzinfo=timezone.utc)


def _match(value="100.00", line_text="") -> MatchRecord:
    return MatchRecord(
        display_name="scan.pdf",
        location="page 1",
        raw_text=f"${value}",
        rule_id="standard",
        value=Decimal(value),
        line_text=line_text,
    )


def _result(matches: list[MatchRecord]) -> PipelineResult:
    return PipelineResult.from_documents(
        [
            DocumentResult(
                display_name="scan.pdf",
                status=Status.OK,
                matches=matches,
                subtotal=sum((m.value for m in matches), Decimal("0")),
            )
        ]
    )


def _sheet(tmp_path, result, name, rules=None):
    path = tmp_path / "report.xlsx"
    save_workbook(build_workbook(result, rules), path)
    return openpyxl.load_workbook(path)[name]


def test_build_workbook_with_no_category_rules_argument_still_produces_details(tmp_path):
    # The default-None path -- every pre-existing single-argument call
    # site keeps compiling and behaving as before.
    ws = _sheet(tmp_path, _result([_match()]), "Details")
    header = [c.value for c in ws[1]]
    row = [c.value for c in ws[2]]

    assert row[header.index("Category")] == "Uncategorized"
    assert row[header.index("Category Review")] == "REVIEW"


def test_details_reports_a_confirmed_category(tmp_path):
    m = _match()
    record_revision(m.category_revisions, "Materials", now=_NOW)
    ws = _sheet(tmp_path, _result([m]), "Details")
    header = [c.value for c in ws[1]]
    row = [c.value for c in ws[2]]

    assert row[header.index("Category")] == "Materials"
    assert row[header.index("Category Review")] is None


def test_details_reports_a_suggested_unconfirmed_category(tmp_path):
    from cost_extractor.category_rules import default_rules

    m = _match(line_text="materials delivered")
    ws = _sheet(tmp_path, _result([m]), "Details", rules=default_rules())
    header = [c.value for c in ws[1]]
    row = [c.value for c in ws[2]]

    assert row[header.index("Category")] == "Materials (suggested, unconfirmed)"
    assert row[header.index("Category Review")] == "REVIEW"


def test_summary_reports_amounts_not_yet_categorized(tmp_path):
    categorized = _match()
    record_revision(categorized.category_revisions, "Materials", now=_NOW)
    uncategorized = _match()
    ws = _sheet(tmp_path, _result([categorized, uncategorized]), "Summary")
    labels = {row[0].value: row[3].value for row in ws.iter_rows()}

    assert labels["Amounts not yet categorized"] == 1


def test_revisions_sheet_gets_a_category_dimension_row(tmp_path):
    m = _match()
    record_revision(m.category_revisions, "Materials", note="from invoice", now=_NOW)
    ws = _sheet(tmp_path, _result([m]), "Revisions")
    header = [c.value for c in ws[1]]
    row = [c.value for c in ws[2]]

    assert row[header.index("Dimension")] == "Category"
    assert row[header.index("Revised From")] == "Uncategorized"
    assert row[header.index("Revised To")] == "Materials"
    assert row[header.index("Note")] == "from invoice"


def test_a_value_revision_row_reads_value_for_dimension(tmp_path):
    m = _match()
    record_revision(m.value_revisions, Decimal("150.00"), now=_NOW)
    ws = _sheet(tmp_path, _result([m]), "Revisions")
    header = [c.value for c in ws[1]]
    row = [c.value for c in ws[2]]

    assert row[header.index("Dimension")] == "Value"


def test_a_second_category_confirmation_shows_two_revision_rows(tmp_path):
    m = _match()
    first = datetime(2026, 9, 3, 10, 14, tzinfo=timezone.utc)
    second = datetime(2026, 9, 3, 10, 22, tzinfo=timezone.utc)
    record_revision(m.category_revisions, "Materials", now=first)
    record_revision(m.category_revisions, "Labor", note="fixed", now=second)
    ws = _sheet(tmp_path, _result([m]), "Revisions")
    header = [c.value for c in ws[1]]
    row1 = [c.value for c in ws[2]]
    row2 = [c.value for c in ws[3]]

    assert row1[header.index("Revised From")] == "Uncategorized"
    assert row1[header.index("Revised To")] == "Materials"
    assert row2[header.index("Revised From")] == "Materials"
    assert row2[header.index("Revised To")] == "Labor"


def test_a_match_with_both_value_and_category_history_shows_value_before_category(tmp_path):
    # Both dimensions on one match, in one Revisions block: Value history
    # must read before Category history, per the spec's ordering rule.
    m = _match()
    record_revision(m.value_revisions, Decimal("150.00"), now=_NOW)
    record_revision(m.category_revisions, "Materials", now=_NOW)
    ws = _sheet(tmp_path, _result([m]), "Revisions")
    header = [c.value for c in ws[1]]
    row1 = [c.value for c in ws[2]]
    row2 = [c.value for c in ws[3]]

    assert row1[header.index("Dimension")] == "Value"
    assert row2[header.index("Dimension")] == "Category"


def test_categories_sheet_lists_confirmed_and_unconfirmed_rows(tmp_path):
    from cost_extractor.category_rules import default_rules

    confirmed = _match(value="100.00")
    record_revision(confirmed.category_revisions, "Materials", now=_NOW)
    unconfirmed = _match(value="50.00", line_text="labor charges")
    ws = _sheet(
        tmp_path, _result([confirmed, unconfirmed]), "Categories", rules=default_rules()
    )
    rows = {
        (row[0], row[1]): (row[2], row[3])
        for row in ws.iter_rows(min_row=2, values_only=True)
    }

    assert rows[("Materials", "Confirmed")] == (100.0, 1)
    assert rows[("Labor", "Unconfirmed")] == (50.0, 1)


def test_categories_sheet_uncategorized_row_omitted_when_none(tmp_path):
    m = _match()
    record_revision(m.category_revisions, "Materials", now=_NOW)
    ws = _sheet(tmp_path, _result([m]), "Categories")
    categories = [row[0] for row in ws.iter_rows(min_row=2, values_only=True)]

    assert "Uncategorized" not in categories


def test_categories_sheet_uncategorized_row_present_when_no_signal(tmp_path):
    m = _match()  # no line_text, no rules passed -- no suggestion possible
    ws = _sheet(tmp_path, _result([m]), "Categories")
    rows = {row[0]: (row[2], row[3]) for row in ws.iter_rows(min_row=2, values_only=True)}

    assert rows["Uncategorized"] == (100.0, 1)


def test_a_confirmed_none_category_lands_in_uncategorized_not_a_crash(tmp_path):
    # Defends against a future writer of category_revisions recording
    # None (the data model allows it; only today's GUI callers forbid
    # it) -- must degrade to the Uncategorized bucket, not crash the
    # whole export via an unsortable None/str mix.
    m = _match()
    record_revision(m.category_revisions, None, now=_NOW)
    ws = _sheet(tmp_path, _result([m]), "Categories")
    rows = {row[0]: (row[2], row[3]) for row in ws.iter_rows(min_row=2, values_only=True)}

    assert rows["Uncategorized"] == (100.0, 1)


def test_categories_sheet_exists_header_only_with_zero_matches(tmp_path):
    ws = _sheet(tmp_path, _result([]), "Categories")

    assert ws.max_row == 1


def test_build_workbook_gains_the_categories_sheet(tmp_path):
    wb = build_workbook(_result([]))

    assert wb.sheetnames == ["Summary", "Details", "Categories", "Revisions"]
