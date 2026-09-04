from decimal import Decimal
from pathlib import Path

import openpyxl

from cost_extractor.extractors.base import Status
from cost_extractor.pipeline import DocumentResult, MatchRecord, PipelineResult
from cost_extractor.report import build_workbook, save_workbook


def _sample_result() -> PipelineResult:
    doc_a = DocumentResult(
        display_name="invoice.docx",
        status=Status.OK,
        matches=[
            MatchRecord(
                display_name="invoice.docx",
                location="paragraph 1",
                raw_text="$1,234.56",
                rule_id="standard",
                value=Decimal("1234.56"),
            ),
            MatchRecord(
                display_name="invoice.docx",
                location="table 1, row 1, col 2",
                raw_text="($500)",
                rule_id="paren_negative",
                value=Decimal("-500"),
            ),
        ],
        subtotal=Decimal("734.56"),
    )
    doc_b = DocumentResult(
        display_name="broken.pdf",
        status=Status.ERROR,
        message="corrupt or unreadable PDF",
        matches=[],
        subtotal=Decimal("0"),
    )
    return PipelineResult(documents=[doc_a, doc_b], grand_total=Decimal("734.56"))


def test_build_workbook_has_details_and_summary_sheets():
    wb = build_workbook(_sample_result())

    assert wb.sheetnames == [
        "Summary",
        "Details",
        "Categories",
        "Revisions",
        "Spend By Month",
    ]


def test_details_sheet_lists_every_match():
    wb = build_workbook(_sample_result())
    ws = wb["Details"]

    header = [c.value for c in ws[1]]
    assert header == [
        "Source File",
        "Location",
        "Matched Text",
        "Rule",
        "Value",
        "Source",
        "Confidence",
        "Review",
        "Read As Text",
        "Category",
        "Category Review",
        "Spend Date",
        "Spend Date Review",
    ]

    # These fixtures come from a text layer, so they carry no score and
    # nothing is flagged. Neither match has a category or a spend date
    # confirmed, and _sample_result()'s matches carry no line_text and
    # its documents no full_text, so neither can be suggested either --
    # they read "Uncategorized"/REVIEW and "Undated"/REVIEW.
    row2 = [c.value for c in ws[2]]
    assert row2 == [
        "invoice.docx", "paragraph 1", "$1,234.56", "standard", 1234.56,
        "text", None, None, None, "Uncategorized", "REVIEW", "Undated", "REVIEW",
    ]

    row3 = [c.value for c in ws[3]]
    assert row3 == [
        "invoice.docx", "table 1, row 1, col 2", "($500)", "paren_negative", -500,
        "text", None, None, None, "Uncategorized", "REVIEW", "Undated", "REVIEW",
    ]


def test_summary_sheet_lists_per_document_totals_and_grand_total():
    wb = build_workbook(_sample_result())
    ws = wb["Summary"]

    header = [c.value for c in ws[1]]
    assert header == [
        "Document", "Status", "Amounts Found", "Subtotal", "Message", "Review",
    ]

    doc_a_row = [c.value for c in ws[2]]
    assert doc_a_row == ["invoice.docx", "OK", 2, 734.56, None, None]

    doc_b_row = [c.value for c in ws[3]]
    assert doc_b_row == [
        "broken.pdf", "ERROR", 0, 0, "corrupt or unreadable PDF", None,
    ]

    grand_total_row = [c.value for c in ws[4]]
    assert grand_total_row[0] == "Grand Total"
    assert grand_total_row[3] == 734.56


def test_saved_workbook_reloads_with_static_values_not_formulas(tmp_path: Path):
    out_path = tmp_path / "report.xlsx"
    save_workbook(build_workbook(_sample_result()), out_path)

    reloaded = openpyxl.load_workbook(out_path, data_only=True)
    ws = reloaded["Summary"]

    grand_total_cell = ws.cell(row=4, column=4).value
    assert grand_total_cell == 734.56
    assert not str(grand_total_cell).startswith("=")
