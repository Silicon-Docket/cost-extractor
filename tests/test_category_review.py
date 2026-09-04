"""Suggesting and confirming a spend category, and its Categories rules."""

import tkinter as tk
from decimal import Decimal

import pytest

from cost_extractor.extractors.base import Status
from cost_extractor.gui import App
from cost_extractor.pipeline import DocumentResult, MatchRecord, PipelineResult


@pytest.fixture
def app():
    try:
        root = tk.Tk()
    except tk.TclError as e:
        pytest.skip(f"Tk unavailable in this environment: {e}")
    root.withdraw()
    application = App(root)
    yield application
    root.destroy()


def _match(raw_text="$100.00", value="100.00", line_text="") -> MatchRecord:
    return MatchRecord(
        display_name="scan.pdf",
        location="page 1",
        raw_text=raw_text,
        rule_id="standard",
        value=Decimal(value),
        line_text=line_text,
    )


def _load(app, matches) -> PipelineResult:
    app.last_result = PipelineResult.from_documents(
        [
            DocumentResult(
                display_name="scan.pdf",
                status=Status.OK,
                matches=matches,
                subtotal=sum((m.value for m in matches), Decimal("0")),
            )
        ]
    )
    return app.last_result


def test_suggest_category_finds_a_match_from_the_line_text(app):
    m = _match(line_text="materials delivered today")
    _load(app, [m])

    assert app.suggest_category(m) == "Materials"


def test_suggest_category_returns_none_with_no_match_on_the_line(app):
    m = _match(line_text="nothing relevant here")
    _load(app, [m])

    assert app.suggest_category(m) is None


def test_suggest_category_is_cached(app):
    m = _match(line_text="materials delivered")
    _load(app, [m])
    first = app.suggest_category(m)

    # Mutate the match's own line_text -- if suggest_category recomputed
    # instead of using the cache, this would change the answer.
    m.line_text = "nothing relevant now"

    assert app.suggest_category(m) == first


def test_confirm_category_records_a_category(app):
    m = _match()
    _load(app, [m])

    error = app.confirm_category(m, "Materials")

    assert error is None
    assert m.effective_category == "Materials"


def test_confirm_category_rejects_an_empty_category(app):
    m = _match()
    _load(app, [m])

    error = app.confirm_category(m, "   ")

    assert error is not None
    assert m.category_revisions == []


def test_a_second_category_confirmation_preserves_the_first_as_history(app):
    m = _match()
    _load(app, [m])

    app.confirm_category(m, "Materials")
    app.confirm_category(m, "Labor", note="reclassified")

    assert [r.value for r in m.category_revisions] == ["Materials", "Labor"]
    assert m.category_revisions[-1].note == "reclassified"
    assert m.effective_category == "Labor"


def test_accept_category_suggestion_confirms_the_suggested_category(app):
    m = _match(line_text="materials delivered today")
    _load(app, [m])

    error = app.accept_category_suggestion(m)

    assert error is None
    assert m.effective_category == "Materials"


def test_accept_category_suggestion_with_no_suggestion_available_is_rejected(app):
    m = _match(line_text="nothing relevant here")
    _load(app, [m])

    error = app.accept_category_suggestion(m)

    assert error is not None
    assert m.category_revisions == []


def test_category_suggestions_cache_is_cleared_between_runs(app, monkeypatch):
    # The exact bug caught late on the sibling spend-over-time branch:
    # a stale id(match)-keyed cache entry from a PRIOR run must not
    # survive into a NEW run's different MatchRecord. Directly plants a
    # stale entry at id(m2) rather than relying on natural CPython
    # id-reuse (which won't happen here since m1 stays alive as a local
    # variable) -- this makes the regression concrete and deterministic:
    # if _run_worker's cache-clear line were removed, this planted value
    # would be returned instead of the correct recomputed None.
    m1 = _match(line_text="materials delivered")
    _load(app, [m1])
    assert app.suggest_category(m1) == "Materials"  # seed the cache

    m2 = _match(raw_text="$200.00", value="200.00", line_text="nothing relevant here")
    doc2 = DocumentResult(
        display_name="scan2.pdf", status=Status.OK, matches=[m2], subtotal=Decimal("200.00"),
    )
    result2 = PipelineResult.from_documents([doc2])

    import cost_extractor.gui as gui_module

    app._category_suggestions[id(m2)] = "Materials"  # plant the stale entry
    monkeypatch.setattr(gui_module, "run_pipeline", lambda *a, **k: result2)
    app._run_worker([], [])

    assert app.suggest_category(m2) is None


def test_add_category_rule_success_adds_rule(app):
    error = app.add_category_rule(r"\bpermits?\b", "Permits")

    assert error is None
    assert any(r.label == "Permits" and not r.built_in for r in app.category_rules)


def test_add_category_rule_invalid_pattern_returns_error_and_does_not_add(app):
    before = len(app.category_rules)

    error = app.add_category_rule(r"\bpermits?\b(", "Broken")

    assert error is not None
    assert len(app.category_rules) == before


def test_remove_category_rule_removes_a_custom_rule(app):
    app.add_category_rule(r"\bpermits?\b", "Permits")
    custom_id = next(r.id for r in app.category_rules if not r.built_in)

    app.remove_category_rule(custom_id)

    assert all(r.built_in for r in app.category_rules)


def test_toggle_category_rule_disables_it(app):
    rule_id = app.category_rules[0].id

    app.toggle_category_rule(rule_id, False)

    assert app.category_rules[0].enabled is False


def test_adding_a_category_rule_invalidates_the_suggestion_cache(app):
    m = _match(line_text="building permit")
    _load(app, [m])

    # Seed the cache BEFORE the matching rule exists, and assert it
    # seeded None -- otherwise a passing test below wouldn't prove the
    # cache was actually cleared rather than never populated.
    assert app.suggest_category(m) is None

    app.add_category_rule(r"\bpermits?\b", "Permits")

    assert app.suggest_category(m) == "Permits"


def test_removing_a_category_rule_invalidates_the_suggestion_cache(app):
    app.add_category_rule(r"\bpermits?\b", "Permits")
    custom_id = next(r.id for r in app.category_rules if not r.built_in)
    m = _match(line_text="building permit")
    _load(app, [m])
    assert app.suggest_category(m) == "Permits"

    app.remove_category_rule(custom_id)

    assert app.suggest_category(m) is None


def test_toggling_a_category_rule_off_invalidates_the_suggestion_cache(app):
    m = _match(line_text="materials delivered")
    _load(app, [m])
    assert app.suggest_category(m) == "Materials"
    materials_id = next(r.id for r in app.category_rules if r.id == "materials")

    app.toggle_category_rule(materials_id, False)

    assert app.suggest_category(m) is None


def test_adding_a_category_rule_refreshes_an_open_category_window(app):
    m = _match(line_text="nothing relevant here")
    _load(app, [m])
    app.open_category_window()
    assert app._category_suggestion_label.cget("text") == "No category suggestion for this line."

    app.add_category_rule(r"\brelevant\b", "Relevant")

    assert app._category_suggestion_label.cget("text") == "Suggested: Relevant"


def test_removing_a_category_rule_refreshes_an_open_category_window(app):
    app.add_category_rule(r"\bpermits?\b", "Permits")
    custom_id = next(r.id for r in app.category_rules if not r.built_in)
    m = _match(line_text="building permit")
    _load(app, [m])
    app.open_category_window()
    assert app._category_suggestion_label.cget("text") == "Suggested: Permits"

    app.remove_category_rule(custom_id)

    assert (
        app._category_suggestion_label.cget("text")
        == "No category suggestion for this line."
    )


def test_toggling_a_category_rule_off_refreshes_an_open_category_window(app):
    m = _match(line_text="materials delivered")
    _load(app, [m])
    app.open_category_window()
    assert app._category_suggestion_label.cget("text") == "Suggested: Materials"
    materials_id = next(r.id for r in app.category_rules if r.id == "materials")

    app.toggle_category_rule(materials_id, False)

    assert (
        app._category_suggestion_label.cget("text")
        == "No category suggestion for this line."
    )


def test_the_category_window_opens_and_shows_the_first_match(app):
    _load(app, [_match(raw_text="$100.00")])

    window = app.open_category_window()

    assert window.winfo_exists()
    assert app.current_category_match().raw_text == "$100.00"


def test_opening_the_category_window_twice_reuses_the_same_window(app):
    _load(app, [_match()])

    first = app.open_category_window()
    second = app.open_category_window()

    assert first is second


def test_the_category_queue_includes_every_match_not_just_ocr(app):
    a = _match(raw_text="$100.00")
    b = _match(raw_text="$200.00")
    _load(app, [a, b])

    assert len(app.category_queue()) == 2


def test_moving_through_the_category_queue_changes_the_shown_match(app):
    _load(app, [_match(raw_text="$100.00"), _match(raw_text="$200.00")])
    app.open_category_window()

    first = app.current_category_match()
    app.next_category_review()

    assert app.current_category_match() is not first


def test_the_category_queue_does_not_run_off_the_end(app):
    _load(app, [_match()])
    app.open_category_window()

    app.next_category_review()
    app.next_category_review()

    assert app.current_category_match() is not None


def test_saving_a_category_through_the_window_advances_the_queue(app):
    _load(app, [_match(raw_text="$100.00"), _match(raw_text="$200.00")])
    app.open_category_window()
    first = app.current_category_match()

    app._category_entry.insert(0, "Materials")
    app._on_save_category()

    assert first.effective_category == "Materials"
    assert app.current_category_match() is not first


def test_the_category_button_is_off_until_a_result_is_loaded(app):
    assert "disabled" in app._category_button.state()

    _load(app, [_match()])
    app._refresh_preview_widget()

    assert "disabled" not in app._category_button.state()


def test_the_categories_panel_lists_the_built_in_rules(app):
    assert len(app._category_rules_container.winfo_children()) == 4  # materials/labor/travel/fees


def test_adding_a_category_rule_through_the_panel_extends_the_checkbox_list(app):
    app._category_pattern_entry.insert(0, r"\bpermits?\b")
    app._category_label_entry.insert(0, "Permits")

    app._on_add_category_rule()

    assert len(app._category_rules_container.winfo_children()) == 5


def test_a_new_run_refreshes_the_open_category_window(app, tmp_path, monkeypatch):
    # Checks the visible caption widget, not just current_category_match():
    # current_category_match() is a pure function of last_result /
    # category_review_index and would already report the new match even
    # without the fix. The actual bug is that the on-screen caption stays
    # stale until _refresh_category_widgets() is called, which is what an
    # operator trusts when clicking "Confirm category".
    import time
    import cost_extractor.gui as gui_module

    m1 = _match(raw_text="$100.00", line_text="first run line")
    _load(app, [m1])
    app.open_category_window()
    assert app.current_category_match() is m1
    assert "first run line" in app._category_caption.cget("text")

    m2 = _match(raw_text="$200.00", value="200.00", line_text="second run line")
    doc2 = DocumentResult(
        display_name="scan2.pdf", status=Status.OK, matches=[m2], subtotal=Decimal("200.00"),
    )
    result2 = PipelineResult.from_documents([doc2])
    monkeypatch.setattr(gui_module, "run_pipeline", lambda *a, **k: result2)

    app.add_paths([tmp_path / "fake.docx"])
    app.start_run()
    deadline = time.time() + 10
    # Wait for _poll_progress itself to finish handling "done" (it clears
    # _worker_thread), not just for last_result to change -- last_result is
    # set by the background thread BEFORE "done" is queued, so waiting on
    # last_result alone races ahead of _poll_progress's refresh calls.
    while app._worker_thread is not None and time.time() < deadline:
        app.root.update()
        time.sleep(0.05)

    assert app.last_result is result2
    assert app.current_category_match() is m2
    caption = app._category_caption.cget("text")
    assert "second run line" in caption
    assert "first run line" not in caption


def test_export_report_passes_the_live_category_rules_through(app, tmp_path):
    import openpyxl

    m = _match(line_text="materials delivered today")
    _load(app, [m])

    path = tmp_path / "report.xlsx"
    error = app.export_report(path)

    assert error is None
    ws = openpyxl.load_workbook(path)["Details"]
    header = [c.value for c in ws[1]]
    row = [c.value for c in ws[2]]
    assert row[header.index("Category")] == "Materials (suggested, unconfirmed)"
