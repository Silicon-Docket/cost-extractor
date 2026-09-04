"""The report has to distinguish a read amount from a guessed one."""

from datetime import datetime, timezone
from decimal import Decimal

import openpyxl

from cost_extractor.extractors.base import Status
from cost_extractor.pipeline import DocumentResult, MatchRecord, PipelineResult
from cost_extractor.report import build_workbook, review_label, save_workbook
from cost_extractor.revisions import record_revision

_FIRST = datetime(2026, 9, 3, 10, 14, tzinfo=timezone.utc)
_SECOND = datetime(2026, 9, 3, 10, 22, tzinfo=timezone.utc)


def _match(value: str, confidence=None, display_name="scan.pdf") -> MatchRecord:
    return MatchRecord(
        display_name=display_name,
        location="page 1",
        raw_text=f"${value}",
        rule_id="standard",
        value=Decimal(value),
        provenance="text" if confidence is None else "ocr",
        confidence=confidence,
    )


def _result() -> PipelineResult:
    return PipelineResult.from_documents(
        [
            DocumentResult(
                display_name="scan.pdf",
                status=Status.OK,
                matches=[
                    _match("100.00"),
                    _match("200.00", confidence=95.0),
                    _match("40.00", confidence=31.0),
                ],
                subtotal=Decimal("340.00"),
            )
        ]
    )


def _sheet(tmp_path, name):
    path = tmp_path / "report.xlsx"
    save_workbook(build_workbook(_result()), path)
    return openpyxl.load_workbook(path)[name]


def test_details_sheet_reports_source_and_confidence(tmp_path):
    ws = _sheet(tmp_path, "Details")

    header = [c.value for c in ws[1]]
    assert "Source" in header
    assert "Confidence" in header


def test_a_read_amount_reports_its_source_with_no_score(tmp_path):
    ws = _sheet(tmp_path, "Details")
    header = [c.value for c in ws[1]]
    row = [c.value for c in ws[2]]

    assert row[header.index("Source")] == "text"
    # A blank cell, not a zero: nothing was guessed, so there is no score.
    assert row[header.index("Confidence")] is None


def test_a_guessed_amount_reports_its_score(tmp_path):
    ws = _sheet(tmp_path, "Details")
    header = [c.value for c in ws[1]]
    row = [c.value for c in ws[3]]

    assert row[header.index("Source")] == "ocr"
    assert row[header.index("Confidence")] == 95.0


def test_a_doubtful_amount_is_marked_for_review(tmp_path):
    ws = _sheet(tmp_path, "Details")
    header = [c.value for c in ws[1]]
    row = [c.value for c in ws[4]]

    assert row[header.index("Review")] == "REVIEW"


def test_a_trusted_amount_is_not_marked_for_review(tmp_path):
    ws = _sheet(tmp_path, "Details")
    header = [c.value for c in ws[1]]
    row = [c.value for c in ws[3]]

    assert row[header.index("Review")] is None


def test_summary_splits_the_total_without_losing_the_combined_one(tmp_path):
    ws = _sheet(tmp_path, "Summary")

    labels = {row[0].value: row[3].value for row in ws.iter_rows()}
    assert labels["Grand Total"] == 340.0
    assert labels["Of which needs review"] == 40.0


def test_summary_flags_a_document_that_needs_review(tmp_path):
    ws = _sheet(tmp_path, "Summary")

    header = [c.value for c in ws[1]]
    row = [c.value for c in ws[2]]
    assert row[header.index("Review")] == "REVIEW"


def _corrected_result() -> PipelineResult:
    read_wrong = _match("440.00", confidence=84.0)
    record_revision(read_wrong.value_revisions, Decimal("940.00"))
    checked = _match("200.00", confidence=95.0)
    record_revision(checked.value_revisions, Decimal("200.00"))
    return PipelineResult.from_documents(
        [
            DocumentResult(
                display_name="scan.pdf",
                status=Status.OK,
                matches=[read_wrong, checked],
                subtotal=Decimal("640.00"),
            )
        ]
    )


def _corrected_sheet(tmp_path, name):
    path = tmp_path / "corrected.xlsx"
    save_workbook(build_workbook(_corrected_result()), path)
    return openpyxl.load_workbook(path)[name]


def test_a_correction_reaches_the_exported_value(tmp_path):
    ws = _corrected_sheet(tmp_path, "Details")
    header = [c.value for c in ws[1]]

    assert [c.value for c in ws[2]][header.index("Value")] == 940.0


def test_the_original_reading_is_still_exported(tmp_path):
    # A correction has to be visible as a correction, not a silent rewrite.
    ws = _corrected_sheet(tmp_path, "Details")
    header = [c.value for c in ws[1]]

    assert [c.value for c in ws[2]][header.index("Read As Text")] == "$440.00"


def test_a_corrected_row_says_so(tmp_path):
    ws = _corrected_sheet(tmp_path, "Details")
    header = [c.value for c in ws[1]]

    assert [c.value for c in ws[2]][header.index("Review")] == "corrected"
    assert [c.value for c in ws[3]][header.index("Review")] == "checked"


def test_the_grand_total_reflects_corrections(tmp_path):
    ws = _corrected_sheet(tmp_path, "Summary")
    labels = {row[0].value: row[3].value for row in ws.iter_rows()}

    assert labels["Grand Total"] == 1140.0  # 940 + 200, not 640


def test_the_summary_reports_what_is_still_unchecked(tmp_path):
    ws = _sheet(tmp_path, "Summary")
    # Column 2 is "Amounts Found" -- this is a count, not money.
    labels = {row[0].value: row[2].value for row in ws.iter_rows()}

    # Two OCR amounts in _result(), neither reviewed.
    assert labels["Guessed amounts not yet checked"] == 2


def test_revisions_sheet_header(tmp_path):
    result = PipelineResult.from_documents([])
    path = tmp_path / "empty.xlsx"
    save_workbook(build_workbook(result), path)
    ws = openpyxl.load_workbook(path)["Revisions"]

    assert [c.value for c in ws[1]] == [
        "Source File", "Location", "Matched Text", "Rule", "Dimension",
        "Revised From", "Revised To", "Timestamp", "Note",
    ]


def test_a_second_correction_shows_two_rows_in_the_revisions_sheet(tmp_path):
    m = _match("440.00", confidence=84.0)
    record_revision(m.value_revisions, Decimal("900.00"), now=_FIRST)
    record_revision(m.value_revisions, Decimal("940.00"), note="fixed typo", now=_SECOND)
    result = PipelineResult.from_documents(
        [
            DocumentResult(
                display_name="scan.pdf", status=Status.OK,
                matches=[m], subtotal=Decimal("440.00"),
            )
        ]
    )
    path = tmp_path / "report.xlsx"
    save_workbook(build_workbook(result), path)
    ws = openpyxl.load_workbook(path)["Revisions"]

    header = [c.value for c in ws[1]]
    row1 = [c.value for c in ws[2]]
    row2 = [c.value for c in ws[3]]

    assert row1[header.index("Revised From")] == 440.0
    assert row1[header.index("Revised To")] == 900.0
    assert row1[header.index("Timestamp")] == "2026-09-03 10:14 UTC"
    assert row1[header.index("Note")] is None
    assert row2[header.index("Revised From")] == 900.0
    assert row2[header.index("Revised To")] == 940.0
    assert row2[header.index("Timestamp")] == "2026-09-03 10:22 UTC"
    assert row2[header.index("Note")] == "fixed typo"


def test_revisions_sheet_disambiguates_matches_sharing_a_location(tmp_path):
    # location is coarse by construction (a whole page/paragraph/image);
    # Matched Text + Rule are what tell two matches on the same page apart.
    a = _match("100.00", confidence=90.0)
    b = _match("200.00", confidence=90.0)
    record_revision(a.value_revisions, Decimal("150.00"), now=_FIRST)
    record_revision(b.value_revisions, Decimal("250.00"), now=_FIRST)
    result = PipelineResult.from_documents(
        [
            DocumentResult(
                display_name="scan.pdf", status=Status.OK,
                matches=[a, b], subtotal=Decimal("300.00"),
            )
        ]
    )
    path = tmp_path / "report.xlsx"
    save_workbook(build_workbook(result), path)
    ws = openpyxl.load_workbook(path)["Revisions"]

    header = [c.value for c in ws[1]]
    matched_text = [
        row[header.index("Matched Text")]
        for row in ws.iter_rows(min_row=2, values_only=True)
    ]

    assert matched_text == ["$100.00", "$200.00"]


def test_revisions_sheet_orders_by_document_then_match_not_by_timestamp(tmp_path):
    # The spec commits to document-then-match-then-revision order, NOT a
    # global chronological sort -- interleave timestamps across two
    # documents to prove it's the former.
    a = _match("100.00", confidence=90.0, display_name="doc_a.pdf")
    record_revision(a.value_revisions, Decimal("150.00"), now=_SECOND)  # later timestamp
    b = _match("200.00", confidence=90.0, display_name="doc_b.pdf")
    record_revision(b.value_revisions, Decimal("250.00"), now=_FIRST)  # earlier timestamp
    result = PipelineResult.from_documents(
        [
            DocumentResult(
                display_name="doc_a.pdf", status=Status.OK,
                matches=[a], subtotal=Decimal("100.00"),
            ),
            DocumentResult(
                display_name="doc_b.pdf", status=Status.OK,
                matches=[b], subtotal=Decimal("200.00"),
            ),
        ]
    )
    path = tmp_path / "report.xlsx"
    save_workbook(build_workbook(result), path)
    ws = openpyxl.load_workbook(path)["Revisions"]

    source_files = [
        row[0] for row in ws.iter_rows(min_row=2, values_only=True)
    ]
    # doc_a comes first (document order), even though doc_b's revision has
    # the earlier timestamp -- proves this is not a global time sort.
    assert source_files == ["doc_a.pdf", "doc_b.pdf"]


def test_review_label_reads_checked_when_a_reverted_correction_lands_back_on_the_original():
    # Intentional current-state semantics: the Revisions sheet has the
    # full history; this label answers "does it differ right now".
    m = _match("440.00", confidence=84.0)
    record_revision(m.value_revisions, Decimal("900.00"), now=_FIRST)
    record_revision(m.value_revisions, Decimal("440.00"), now=_SECOND)  # reverted

    assert review_label(m) == "checked"


def test_a_match_with_no_revisions_has_no_revisions_sheet_rows(tmp_path):
    result = PipelineResult.from_documents(
        [
            DocumentResult(
                display_name="scan.pdf", status=Status.OK,
                matches=[_match("100.00")], subtotal=Decimal("100.00"),
            )
        ]
    )
    path = tmp_path / "report.xlsx"
    save_workbook(build_workbook(result), path)
    ws = openpyxl.load_workbook(path)["Revisions"]

    assert ws.max_row == 1  # header only
