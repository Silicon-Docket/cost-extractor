"""Suggesting and confirming a spend date, and its Date Formats rules."""

import tkinter as tk
from datetime import date
from decimal import Decimal

import pytest

from cost_extractor.extractors.base import Status
from cost_extractor.gui import App
from cost_extractor.pipeline import DocumentResult, MatchRecord, PipelineResult


@pytest.fixture
def app():
    try:
        root = tk.Tk()
    except tk.TclError as e:
        pytest.skip(f"Tk unavailable in this environment: {e}")
    root.withdraw()
    application = App(root)
    yield application
    root.destroy()


def _match(raw_text="$100.00", value="100.00", doc_offset=0) -> MatchRecord:
    return MatchRecord(
        display_name="scan.pdf",
        location="page 1",
        raw_text=raw_text,
        rule_id="standard",
        value=Decimal(value),
        doc_offset=doc_offset,
    )


def _load(app, matches, full_text=""):
    """Loads a result the way a real run does, INCLUDING the
    match -> document map _run_worker builds -- a test that skips this
    step (by setting app.last_result directly) leaves _document_for with
    nothing to look up."""
    doc = DocumentResult(
        display_name="scan.pdf",
        status=Status.OK,
        matches=matches,
        subtotal=sum((m.value for m in matches), Decimal("0")),
        full_text=full_text,
    )
    app.last_result = PipelineResult.from_documents([doc])
    app._match_documents = {id(m): doc for m in matches}
    return app.last_result


def test_document_for_finds_the_owning_document(app):
    m = _match()
    result = _load(app, [m])

    assert app._document_for(m) is result.documents[0]


def test_suggest_spend_date_finds_the_nearest_date_in_the_document(app):
    full_text = "Invoice dated 06/14/2026.\n\nAmount: $100.00 due."
    m = _match(doc_offset=full_text.index("$100.00"))
    _load(app, [m], full_text=full_text)

    assert app.suggest_spend_date(m) == date(2026, 6, 14)


def test_suggest_spend_date_returns_none_with_no_dates_in_the_document(app):
    full_text = "No dates anywhere in this text, just $100.00."
    m = _match(doc_offset=full_text.index("$100.00"))
    _load(app, [m], full_text=full_text)

    assert app.suggest_spend_date(m) is None


def test_suggest_spend_date_is_cached(app):
    full_text = "Dated 06/14/2026, amount $100.00."
    m = _match(doc_offset=full_text.index("$100.00"))
    _load(app, [m], full_text=full_text)
    first = app.suggest_spend_date(m)

    # Swap the document out from under the cache -- if suggest_spend_date
    # recomputed instead of using the cache, this would change the answer.
    app._match_documents[id(m)] = DocumentResult(
        display_name="scan.pdf", status=Status.OK, matches=[m], full_text="nothing here"
    )

    assert app.suggest_spend_date(m) == first


def test_confirm_spend_date_records_a_parsed_date(app):
    m = _match()
    _load(app, [m])

    error = app.confirm_spend_date(m, "06/14/2026")

    assert error is None
    assert m.effective_spend_date == date(2026, 6, 14)


def test_a_second_spend_date_confirmation_preserves_the_first_as_history(app):
    m = _match()
    _load(app, [m])

    app.confirm_spend_date(m, "06/01/2026")
    app.confirm_spend_date(m, "06/14/2026", note="fixed typo")

    assert [r.value for r in m.spend_date_revisions] == [
        date(2026, 6, 1),
        date(2026, 6, 14),
    ]
    assert m.spend_date_revisions[-1].note == "fixed typo"
    assert m.effective_spend_date == date(2026, 6, 14)


def test_confirm_spend_date_rejects_unparseable_text(app):
    m = _match()
    _load(app, [m])

    error = app.confirm_spend_date(m, "not a date")

    assert error is not None
    assert m.spend_date_revisions == []


def test_accept_date_suggestion_confirms_the_suggested_date(app):
    full_text = "Dated 06/14/2026, amount $100.00."
    m = _match(doc_offset=full_text.index("$100.00"))
    _load(app, [m], full_text=full_text)

    error = app.accept_date_suggestion(m)

    assert error is None
    assert m.effective_spend_date == date(2026, 6, 14)


def test_accept_date_suggestion_with_no_suggestion_available_is_rejected(app):
    full_text = "No dates here, just $100.00."
    m = _match(doc_offset=full_text.index("$100.00"))
    _load(app, [m], full_text=full_text)

    error = app.accept_date_suggestion(m)

    assert error is not None
    assert m.spend_date_revisions == []


def test_confirm_no_date_records_a_none_valued_revision(app):
    m = _match()
    _load(app, [m])

    app.confirm_no_date(m)

    assert m.spend_date_reviewed is True
    assert m.effective_spend_date is None


def test_confirm_no_date_default_note(app):
    m = _match()
    _load(app, [m])

    app.confirm_no_date(m)

    assert m.spend_date_revisions[-1].note == "confirmed no associated date"


def test_add_date_rule_success_adds_rule(app):
    error = app.add_date_rule(r"(?P<year>\d{4})-(?P<month>\d{2})-(?P<day>\d{2})", "ISO")

    assert error is None
    assert any(r.label == "ISO" and not r.built_in for r in app.date_rules)


def test_add_date_rule_invalid_pattern_returns_error_and_does_not_add(app):
    before = len(app.date_rules)

    error = app.add_date_rule(r"(?P<year>\d{4}", "Broken")

    assert error is not None
    assert len(app.date_rules) == before


def test_remove_date_rule_removes_a_custom_rule(app):
    app.add_date_rule(r"(?P<year>\d{4})-(?P<month>\d{2})-(?P<day>\d{2})", "ISO")
    custom_id = next(r.id for r in app.date_rules if not r.built_in)

    app.remove_date_rule(custom_id)

    assert all(r.built_in for r in app.date_rules)


def test_remove_date_rule_refuses_to_remove_a_built_in_rule(app):
    built_in_id = next(r.id for r in app.date_rules if r.built_in)
    before = len(app.date_rules)

    app.remove_date_rule(built_in_id)

    assert len(app.date_rules) == before


def test_toggle_date_rule_disables_it(app):
    rule_id = app.date_rules[0].id

    app.toggle_date_rule(rule_id, False)

    assert app.date_rules[0].enabled is False


def test_adding_a_date_rule_invalidates_the_suggestion_cache(app):
    full_text = "Reference 2026-06-14, amount $100.00."
    m = _match(doc_offset=full_text.index("$100.00"))
    _load(app, [m], full_text=full_text)

    # Seed the cache BEFORE the matching rule exists, and assert it
    # seeded None -- otherwise a passing test below wouldn't prove the
    # cache was actually cleared rather than never populated.
    assert app.suggest_spend_date(m) is None

    app.add_date_rule(r"(?P<year>\d{4})-(?P<month>\d{2})-(?P<day>\d{2})", "ISO")

    assert app.suggest_spend_date(m) == date(2026, 6, 14)


def test_removing_a_date_rule_invalidates_the_suggestion_cache(app):
    app.add_date_rule(r"(?P<year>\d{4})-(?P<month>\d{2})-(?P<day>\d{2})", "ISO")
    custom_id = next(r.id for r in app.date_rules if not r.built_in)
    full_text = "Reference 2026-06-14, amount $100.00."
    m = _match(doc_offset=full_text.index("$100.00"))
    _load(app, [m], full_text=full_text)
    assert app.suggest_spend_date(m) == date(2026, 6, 14)

    app.remove_date_rule(custom_id)

    assert app.suggest_spend_date(m) is None


def test_toggling_a_date_rule_off_invalidates_the_suggestion_cache(app):
    full_text = "Dated 06/14/2026, amount $100.00."
    m = _match(doc_offset=full_text.index("$100.00"))
    _load(app, [m], full_text=full_text)
    assert app.suggest_spend_date(m) == date(2026, 6, 14)
    builtin_id = app.date_rules[0].id

    app.toggle_date_rule(builtin_id, False)

    assert app.suggest_spend_date(m) is None


def test_the_spend_date_window_opens_and_shows_the_first_match(app):
    _load(app, [_match(raw_text="$100.00")])

    window = app.open_spend_date_window()

    assert window.winfo_exists()
    assert app.current_spend_date_match().raw_text == "$100.00"


def test_opening_the_spend_date_window_twice_reuses_the_same_window(app):
    _load(app, [_match()])

    first = app.open_spend_date_window()
    second = app.open_spend_date_window()

    assert first is second


def test_the_spend_date_queue_includes_every_match_not_just_ocr(app):
    a = _match(raw_text="$100.00")
    b = _match(raw_text="$200.00")
    _load(app, [a, b])

    assert len(app.spend_date_queue()) == 2


def test_moving_through_the_spend_date_queue_changes_the_shown_match(app):
    _load(app, [_match(raw_text="$100.00"), _match(raw_text="$200.00")])
    app.open_spend_date_window()

    first = app.current_spend_date_match()
    app.next_spend_date_review()

    assert app.current_spend_date_match() is not first


def test_the_spend_date_queue_does_not_run_off_the_end(app):
    _load(app, [_match()])
    app.open_spend_date_window()

    app.next_spend_date_review()
    app.next_spend_date_review()

    assert app.current_spend_date_match() is not None


def test_saving_a_spend_date_through_the_window_advances_the_queue(app):
    _load(app, [_match(raw_text="$100.00"), _match(raw_text="$200.00")])
    app.open_spend_date_window()
    first = app.current_spend_date_match()

    app._spend_date_entry.insert(0, "06/14/2026")
    app._on_save_spend_date()

    assert first.effective_spend_date == date(2026, 6, 14)
    assert app.current_spend_date_match() is not first


def test_confirm_no_date_through_the_window_advances_the_queue(app):
    _load(app, [_match(raw_text="$100.00"), _match(raw_text="$200.00")])
    app.open_spend_date_window()
    first = app.current_spend_date_match()

    app._on_confirm_no_date()

    assert first.spend_date_reviewed is True
    assert first.effective_spend_date is None
    assert app.current_spend_date_match() is not first


def test_the_spend_date_button_is_off_until_a_result_is_loaded(app):
    assert "disabled" in app._spend_date_button.state()

    _load(app, [_match()])
    app._refresh_preview_widget()

    assert "disabled" not in app._spend_date_button.state()


def test_the_date_rules_panel_lists_the_built_in_rule(app):
    assert len(app._date_rules_container.winfo_children()) == 1  # numeric_date


def test_adding_a_date_rule_through_the_panel_extends_the_checkbox_list(app):
    app._date_pattern_entry.insert(
        0, r"(?P<year>\d{4})-(?P<month>\d{2})-(?P<day>\d{2})"
    )
    app._date_label_entry.insert(0, "ISO")

    app._on_add_date_rule()

    assert len(app._date_rules_container.winfo_children()) == 2


def test_a_confirmed_dates_prefill_can_be_resaved_without_editing(app):
    m = _match()
    _load(app, [m])
    app.confirm_spend_date(m, "06/14/2026")
    app.open_spend_date_window()  # triggers the prefill via _refresh_spend_date_widgets

    prefilled = app._spend_date_entry.get()
    error = app.confirm_spend_date(m, prefilled)

    assert error is None


def test_date_suggestions_cache_is_cleared_between_runs(app, monkeypatch):
    full_text_1 = "Dated 06/14/2026, amount $100.00."
    m1 = _match(doc_offset=full_text_1.index("$100.00"))
    _load(app, [m1], full_text=full_text_1)
    assert app.suggest_spend_date(m1) == date(2026, 6, 14)  # seed the cache

    # Simulate a second run producing a brand-new match, going through the
    # real _run_worker path (not the _load test shortcut) so the cache
    # invalidation this fix adds is actually exercised.
    full_text_2 = "No dates in this one, amount $200.00."
    m2 = _match(raw_text="$200.00", value="200.00", doc_offset=full_text_2.index("$200.00"))
    doc2 = DocumentResult(
        display_name="scan2.pdf",
        status=Status.OK,
        matches=[m2],
        subtotal=Decimal("200.00"),
        full_text=full_text_2,
    )
    result2 = PipelineResult.from_documents([doc2])

    # Plant a stale entry at m2's own id -- this is what id() reuse
    # produces in a real second run (a freed match's id handed to a new
    # MatchRecord). Without the fix's cache.clear(), suggest_spend_date
    # would return this leaked value instead of recomputing.
    app._date_suggestions[id(m2)] = date(2026, 1, 1)

    import cost_extractor.gui as gui_module

    monkeypatch.setattr(gui_module, "run_pipeline", lambda *args, **kwargs: result2)
    app._run_worker([], [])

    assert app.suggest_spend_date(m2) is None  # not the stale date planted above


def test_export_report_passes_the_live_date_rules_through(app, tmp_path):
    # export_report must hand build_workbook the app's actual date_rules,
    # not the default None -- otherwise this would read "Undated" instead
    # of a suggestion, same as if no rules were ever passed at all.
    import openpyxl

    full_text = "Dated 06/14/2026, amount $100.00."
    m = _match(doc_offset=full_text.index("$100.00"))
    _load(app, [m], full_text=full_text)

    path = tmp_path / "report.xlsx"
    error = app.export_report(path)

    assert error is None
    ws = openpyxl.load_workbook(path)["Details"]
    header = [c.value for c in ws[1]]
    row = [c.value for c in ws[2]]
    assert row[header.index("Spend Date")] == "2026-06-14 (suggested, unconfirmed)"
